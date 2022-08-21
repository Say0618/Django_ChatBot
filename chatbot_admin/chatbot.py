from ast import Not
import openpyxl
from os.path import exists
# from datetime import datetime
import datetime
import base64
import json
from PIL import Image 
import sys
import time

from .calc_algo_v6 import *
import os
import pathlib

from .models import MasterSheet
from .models import ReadSheet
from .models import InterpretationSheet
from .models import Images_Bot
from .models import Database_Excel

# if ReadSheet.objects.count() > 0:
#     if ReadSheet.objects.filter(status=1).count() > 0:
#         read_path = os.getcwd() + "/media/read_sheets/" + ReadSheet.objects.filter(status=1).get().filename()

# write_path = os.getcwd() + "/media/write_sheets/write.xlsx"

# if MasterSheet.objects.count() > 0:
#     if MasterSheet.objects.filter(status=1).count() > 0:
#         master_path = os.getcwd() + "/media/master_sheets/" + MasterSheet.objects.filter(status=1).get().filename()

# if InterpretationSheet.objects.count() > 0:
#     if InterpretationSheet.objects.filter(status=1).count() > 0:
#         interpretation_path = os.getcwd() + "/media/interpretation_sheets/" + InterpretationSheet.objects.filter(status=1).get().filename()



dataset = []
def getData():

    wb = openpyxl.load_workbook(read_path)
    ws = wb.active
    # sheet_names = wb.sheetnames
    rows_cnt = ws.max_row
    cols_cnt = ws.max_column

    start_row = 0 #count real data starts
    flag = False
    for r in range(1, rows_cnt):
        if flag == True:
            break
        for c in range(1, cols_cnt):
            if ws.cell(row=r, column=c).value == "Q.No":
                start_row = r + 1
                flag = True
                break
    flag = False
    for r in range(start_row, rows_cnt+1):
        subset = {}
        ans_set = {}
        if (ws.cell(row=r, column=6).value) is None:
            break
        if ws.cell(row=r, column=2).value:
            flag = False
            subset['order'] = ws.cell(row=r, column=2).value

            subset['question'] = ws.cell(row=r, column=4).value
            subset['type'] = ws.cell(row=r, column=5).value
            question_media = ws.cell(row=r, column=3).value
            #date of birth check
            dob = ws.cell(row=r, column=9).value
            if dob != None:
                subset['dob'] = 1
            #end of the chat
            # end = ws.cell(row=r, column=9).value
            # if end != None:
            #     subset['end'] = True

            if question_media != None:
                if ws.cell(row=r, column=5).value == "Q. picture multiple response" or ws.cell(row=r, column=5).value == "Q. picture single response":
                    question_media = getImageData(question_media)
                subset['question_media'] = question_media

            answer = ws.cell(row=r, column=7).value
            if ws.cell(row=r, column=5).value == "Single picture response" or ws.cell(row=r, column=5).value == "Multiple picture response":
                img_path = ws.cell(row=r, column=7).value.split('>')[0]
                img_name = ws.cell(row=r, column=7).value.split('>')[1]
                flag = True
                answer = getImageData(img_path)
                next = ws.cell(row=r, column=8).value
                # ans_set[next] = answer + ">" + img_name
                ans_set[answer + ">" + img_name] = next
                subset['ans_set'] = ans_set
            else:
                next = ws.cell(row=r, column=8).value
                ans_set[answer] = next
                subset['ans_set'] = ans_set
            dataset.append(subset)
        else:
            subset = dataset[len(dataset)-1]
            if flag == True:
                img_path = ws.cell(row=r, column=7).value.split('>')[0]
                img_name = ws.cell(row=r, column=7).value.split('>')[1]
                answer = getImageData(img_path)
                subset['ans_set'][answer + ">" + img_name] = ws.cell(row=r, column=8).value

            else:
                subset['ans_set'][ws.cell(row=r, column=7).value] = ws.cell(row=r, column=8).value

    
    return dataset

def getImageData(img_name):
    basewidth = 1600
    default_path = os.getcwd() + '/media/attachments/images/'
    img_path = Images_Bot.objects.filter(name=img_name).filter(status=1).get().filename()

    image = Image.open(default_path + img_path)
    wpercent = (basewidth / float(image.size[0]))
    hsize = int((float(image.size[1]) * float(wpercent)))
    image = image.resize((basewidth, hsize), Image.ANTIALIAS)
    image.save(default_path + img_path)

    f=open(default_path + img_path, "rb")
    data = base64.b64encode(f.read()).decode('utf-8')

    return data

def getFirstQuestion():
    return dataset[0]


def getAlpha():
    wb = openpyxl.load_workbook(read_path)
    ws = wb.active
    rows_cnt = ws.max_row
    cols_cnt = ws.max_column

    start_row = 0 #count real data starts
    flag = False
    for r in range(1, rows_cnt):
        if flag == True:
            break
        for c in range(1, cols_cnt):
            if ws.cell(row=r, column=c).value == "Q.No":
                start_row = r + 1
                flag = True
                break

    alpha_set = []
    flag = False
    for r in range(start_row, rows_cnt+1):
        subset = {}
        a_no = []
        if (ws.cell(row=r, column=6).value) is None:
            break
        if ws.cell(row=r, column=2).value:
            flag = False
            subset['order'] = ws.cell(row=r, column=2).value
            if ws.cell(row=r, column=5).value == 'Single picture response' or ws.cell(row=r, column=5).value == 'Multiple picture response' or ws.cell(row=r, column=5).value == 'Single media response' or ws.cell(row=r, column=5).value == 'Multiple media response':
                pick = ws.cell(row=r, column=7).value.split('>')[1]
                a_no.append(pick)
                flag = True
            else:
                a_no.append(ws.cell(row=r, column=7).value)
            subset['a_no'] = a_no
            alpha_set.append(subset)
        else:
            subset = alpha_set[len(alpha_set)-1]
            if flag == True:
                pick = ws.cell(row=r, column=7).value.split('>')[1]
                subset['a_no'].append(pick)
            else:
                subset['a_no'].append(ws.cell(row=r, column=7).value)

    return alpha_set
    print("alpha set is ....",alpha_set)

def writeExcel(write_data_set):
    wb1 = openpyxl.load_workbook(read_path)
    ws1 = wb1.active
    rows_cnt = ws1.max_row
    cols_cnt = ws1.max_column

    start_row = 0 #count real data starts
    flag = False
    for r in range(1, rows_cnt):
        if flag == True:
            break
        for c in range(1, cols_cnt):
            if ws1.cell(row=r, column=c).value == "Q.No":
                start_row = r + 1
                flag = True
                break


    ans = write_data_set[0]
    alpha_set = getAlpha()

    for obj_ans in ans:
        break_flag = False
        for obj_alpha in alpha_set:
            ano_len = len(obj_alpha['a_no'])
            if obj_ans['order'] == obj_alpha['order']:
                if break_flag:
                    break
                for i in range(ano_len):
                    if isinstance(obj_alpha['a_no'][i], str):
                        if obj_alpha['a_no'][i].startswith(obj_ans['ans']):
                            obj_ans['ans'] = i + 1
                            break_flag = True
                            break


    start_time = write_data_set[1].split(' ')[1]
    date = write_data_set[1].split(' ')[0]

    now = datetime.datetime.now()
    last_time = now.strftime("%d/%m/%Y %H:%M:%S")
    last_time = last_time.split(' ')[1]
    
    file_exists = exists(write_path)
    if file_exists == True:
        wb = openpyxl.load_workbook(write_path)
        ws = wb.active

        max_col = ws.max_column
        max_row = ws.max_row
        start_point = max_col + 1

        c1= ws.cell(row=1, column=start_point)
        c1.value = str(start_point-4)

        c1= ws.cell(row=2, column=start_point)
        c1.value = "Code " + str(start_point-4)

        c1= ws.cell(row=3, column=start_point)
        c1.value = "usename " + str(start_point-4)  # type: ignore

        c1= ws.cell(row=4, column=start_point)
        c1.value = start_time

        c1= ws.cell(row=5, column=start_point)
        c1.value = last_time  # type: ignore

        c1= ws.cell(row=6, column=start_point)
        c1.value = date

        increment=0
        for r in range(start_row, rows_cnt+1):
            if (ws1.cell(row=r, column=6).value) is None:
                break
            if ws1.cell(row=r, column=5).value == "No response":
                ws.cell(row=7+increment, column=start_point).value = "A"
            increment+=1

        for obj in ans:
            if isinstance(obj['ans'], str):
                for r in range(7, max_row+1):
                    if ws.cell(row=r, column=3).value == obj['order']:
                        ws.cell(row=r, column=start_point).value = obj['ans']
                        break
            else:
                for r in range(7, max_row+1):
                    if ws.cell(row=r, column=3).value == obj['order']:
                        ws.cell(row=r+obj['ans']-1, column=start_point).value = numToAlpha(obj['ans'])
                        break

        wb.save(write_path)
            

    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        
        c1= ws.cell(row=1, column=1)
        c1.value = "Sl No"  # type: ignore
        c1= ws.cell(row=1, column=2)
        c1.value = "Category"   # type: ignore
        c1= ws.cell(row=1, column=3)
        c1.value = "conv_num"
        c1= ws.cell(row=1, column=4)
        c1.value = "Sub-features"
        c1= ws.cell(row=1, column=5)
        c1.value = "1" 

        c1= ws.cell(row=2, column=3)
        c1.value = "Ideal code"
        c1= ws.cell(row=2, column=5)
        c1.value = "Code 1"

        c1= ws.cell(row=3, column=3)
        c1.value = "usename" 
        c1= ws.cell(row=3, column=5)
        c1.value = "usename 1"

        c1= ws.cell(row=4, column=3)
        c1.value = "start time"
        c1= ws.cell(row=4, column=5)
        c1.value = start_time

        c1= ws.cell(row=5, column=3)
        c1.value = "end time"
        c1= ws.cell(row=5, column=5)
        c1.value = last_time

        c1= ws.cell(row=6, column=3)
        c1.value = "date"
        c1= ws.cell(row=6, column=5)
        c1.value = date

        increment=0
        for r in range(start_row, rows_cnt+1):
            if (ws1.cell(row=r, column=6).value) is None:
                break
            if ws1.cell(row=r, column=5).value == "No response":
                ws.cell(row=7+increment, column=5).value = "A"

            if ws1.cell(row=r,column=1).value != None:
                ws.cell(row=7+increment,column=2).value = str(ws1.cell(row=r,column=1).value)
            
            if ws1.cell(row=r,column=2).value != None:
                ws.cell(row=7+increment,column=3).value = str(ws1.cell(row=r,column=2).value)

            ws.cell(row=7+increment,column=4).value=ws1.cell(row=r,column=6).value

            ws.cell(row=7+increment,column=1).value = increment + 1
            
            increment+=1
        wb.save(write_path)

        wb = openpyxl.load_workbook(write_path)
        ws = wb.active
        max_row = ws.max_row
        for obj in ans:
            if isinstance(obj['ans'], str):
                for r in range(1, max_row+1):
                    if ws.cell(row=r, column=3).value == obj['order']:
                        ws.cell(row=r, column=5).value = obj['ans']
                        break
            else:
                for r in range(1, max_row+1):
                    if ws.cell(row=r, column=3).value == obj['order']:
                        ws.cell(row=r+obj['ans']-1, column=5).value = numToAlpha(obj['ans'])
                        break

        wb.save(write_path)

    data_science()

def data_science():
    process_master_file(master_path)

    process_case_file(write_path)
    do_calculation()

    output_file_name = os.getcwd() + '/media/aa_outputs/aaOutputSheet.xlsx'
    
    global aa
    aa = output_file_name
    output_result_to_excel(output_file_name)

def getFeedback(mode):
    time.sleep(7)
    feedbacks = []
    aa_path = aa
    wb1 = openpyxl.load_workbook(aa_path)
    ws1 = wb1.active
    rows_cnt = ws1.max_row
    cols_cnt = ws1.max_column

    wb2 = openpyxl.load_workbook(interpretation_path)
    ws2_user = wb2['User']
    ws2_dic = wb2['Code Dictionary']

    Score = ws1.cell(row=14, column=cols_cnt).value
    Index = int(ws1.cell(row=13, column=cols_cnt).value)  # type: ignore
    print(Index)

    if Index < 0:
        Index = 1
    if Index > 4:
        Index = 5
#when choose test
    if mode == "User":
        if Score == None:
            elem1 = ws2_user.cell(row=3+Index, column=3).value
            elem2 = ws2_user.cell(row=3+Index, column=5).value
            elem3 = ws2_user.cell(row=3+Index, column=6).value
            elem4 = ws2_user.cell(row=3+Index, column=7).value
            feedbacks.extend([elem1, elem2, elem3, elem4])
        else:
            if Score == "A":
                elem1 = ws2_user.cell(row=2, column=3).value
                elem2 = ws2_user.cell(row=2, column=5).value
                elem3 = ws2_user.cell(row=2, column=6).value
                elem4 = ws2_user.cell(row=2, column=7).value
                feedbacks.extend([elem1, elem2, elem3, elem4])
            elif Score == "B":
                elem1 = ws2_user.cell(row=3, column=3).value
                elem2 = ws2_user.cell(row=3, column=5).value
                elem3 = ws2_user.cell(row=3, column=6).value
                elem4 = ws2_user.cell(row=3, column=7).value
                feedbacks.extend([elem1, elem2, elem3, elem4])

        print(feedbacks)
        # print(len(feedbacks))
        return feedbacks
    else:
        top = []
        top1 = ws1.cell(row=3, column=cols_cnt).value
        top2 = ws1.cell(row=5, column=cols_cnt).value
        top3 = ws1.cell(row=7, column=cols_cnt).value
        top.extend([top1, top2, top3])

        predictions = []
        for item in top:
            for r in range(1, rows_cnt):
                if ws2_dic.cell(row=r, column=1).value == item:
                    predictions.append(ws2_dic.cell(row=r, column=2).value)
        
        for x in predictions:
            feedbacks.append(x)
        
        if Score == None:
            elem1 = ws2_user.cell(row=3+Index, column=4).value
            elem2 = ws2_user.cell(row=3+Index, column=5).value
            elem3 = ws2_user.cell(row=3+Index, column=6).value
            elem4 = ws2_user.cell(row=3+Index, column=7).value
            feedbacks.extend([elem1, elem2, elem3, elem4])
        else:
            if Score == "A":
                elem1 = ws2_user.cell(row=2, column=4).value
                elem2 = ws2_user.cell(row=2, column=5).value
                elem3 = ws2_user.cell(row=2, column=6).value
                elem4 = ws2_user.cell(row=2, column=7).value
                feedbacks.extend([elem1, elem2, elem3, elem4])
            elif Score == "B":
                elem1 = ws2_user.cell(row=3, column=4).value
                elem2 = ws2_user.cell(row=3, column=5).value
                elem3 = ws2_user.cell(row=3, column=6).value
                elem4 = ws2_user.cell(row=3, column=7).value
                feedbacks.extend([elem1, elem2, elem3, elem4])
       

        print(feedbacks)
        # print(len(feedbacks))
        return feedbacks
    
def getDatabase(filename):
    base_path = os.getcwd() + '/media/attachments/database/'

    filename = Database_Excel.objects.filter(status=1).filter(name=filename).get().filename()

    database_path = base_path + filename

    wb = openpyxl.load_workbook(database_path)
    ws = wb.active
    rows_cnt = ws.max_row
    
    result = []
    for r in range(2, rows_cnt + 1):
        item = ws.cell(row=r, column=1).value
        result.append(item)
    
    return result


def numToAlpha(param):
    if param == 1:
        return "A"
    if param == 2:
        return "B"
    if param == 3:
        return "C"
    if param == 4:
        return "D"
    if param == 5:
        return "E"
    if param == 6:
        return "F"
    if param == 7:
        return "G"
    if param == 8:
        return "H"
    if param == 9:
        return "I"
    if param == 10:
        return "J"
    if param == 11:
        return "K"
    if param == 12:
        return "L"
    if param == 13:
        return "M"
    if param == 14:
        return "N"
    if param == 15:
        return "O"
    if param == 16:
        return "P"
    if param == 17:
        return "Q"
    if param == 18:
        return "R"
    if param == 19:
        return "S"
    if param == 20:
        return "T"
    if param == 21:
        return "U"
    if param == 22:
        return "V"
    if param == 23:
        return "W"
    if param == 24:
        return "X"
    if param == 25:
        return "Y"
    if param == 26:
        return "Z"
