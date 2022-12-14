import pkgutil
from django.shortcuts import render, redirect

# Create your views here.
from django.http import HttpResponse
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required
from django.views.decorators.csrf import csrf_exempt, csrf_protect
from django.http import JsonResponse
import datetime
from django.core import serializers
from django.contrib.auth.hashers import make_password, check_password


# import zipfile
from zipfile import ZipFile
import json

# import StringIO
from io import BytesIO
# try:
#     from StringIO import StringIO
# except ImportError:
#     from io import StringIO
import os
import mimetypes
from django.conf import settings as conf_settings
import openpyxl
from os.path import exists

from django.contrib.auth.models import User
from .models import MasterSheet
from .models import ReadSheet
from .models import InterpretationSheet
from .models import Images_Bot
from .models import Database_Excel
from .models import Settings
from .models import Settings_Image
from .models import TestSheet

from .forms import MasterSheetForm
from .forms import ReadSheetForm
from .forms import InterpretationSheetForm
from .forms import Images_BotForm
from .forms import Database_ExcelForm
from .forms import SettingsForm
from .forms import TestSheetForm
#chat bot 
from .chatbot import *

# dataset = []

@login_required
def index(request):
    if request.user.is_authenticated:
        if not request.user.is_superuser:
            return redirect('userIndex')
        else:
            return render(request, 'index.html')
    else:
        return render(request, 'registration/login.html')

@csrf_protect
def login_attempt(request):
    username = request.POST['username']
    password = request.POST['password']

    user = authenticate(request, username=username, password=password)
    
    if user is not None and user.is_active and user.is_superuser:
        login(request, user)
        return redirect(conf_settings.LOGIN_REDIRECT_URL)
    else:
        return render(request, 'registration/login.html', {'msg': 'Login failed!'})    

def logout_attempt(request):
    logout(request)
    return redirect(settings.LOGOUT_REDIRECT_URL)

@login_required
def settings(request):
    terms = ''
    if Settings.objects.filter(type='terms').count() > 0:
        terms = Settings.objects.filter(type='terms').get().content
    cookie = ''
    if Settings.objects.filter(type='cookie').count() > 0:
        cookie = Settings.objects.filter(type='cookie').get().content
    title = ''
    if Settings.objects.filter(type='title').count() > 0:
        title = Settings.objects.filter(type='title').get().content
    welcome = ''
    if Settings.objects.filter(type='welcome').count() > 0:
        welcome = Settings.objects.filter(type='welcome').get().content
    
    logo = ''
    if Settings_Image.objects.filter(type='logo').count() > 0:
        logo = Settings_Image.objects.filter(type='logo').get().name
    login = ''
    if Settings_Image.objects.filter(type='login').count() > 0:
        login = Settings_Image.objects.filter(type='login').get().name
    favicon = ''
    if Settings_Image.objects.filter(type='favicon').count() > 0:
        favicon = Settings_Image.objects.filter(type='favicon').get().name

    return render(request, 'settings.html', {
        'terms': terms,
        'cookie': cookie,
        'title': title,
        'welcome': welcome,
        'logo': logo,
        'login': login,
        'favicon': favicon
    })

def terms_save(request):
    if request.method == 'POST':
        content = request.POST['content']

        if Settings.objects.filter(type='terms').count() > 0:
            terms = Settings.objects.filter(type='terms').get()
            terms.content = content
            terms.save()
        else:
            newModel = Settings.objects.create(content=content, type='terms')
            newModel.save()
        
        return JsonResponse({
            'msg': 'success'
        })

def cookie_save(request):
    if request.method == 'POST':
        content = request.POST['content']

        if Settings.objects.filter(type='cookie').count() > 0:
            cookie = Settings.objects.filter(type='cookie').get()
            cookie.content = content
            cookie.save()
        else:
            newModel = Settings.objects.create(content=content, type='cookie')
            newModel.save()
        
        return JsonResponse({
            'msg': 'success'
        })

def title_save(request):
    if request.method == 'POST':
        content = request.POST['content']

        if Settings.objects.filter(type='title').count() > 0:
            terms = Settings.objects.filter(type='title').get()
            terms.content = content
            terms.save()
        else:
            newModel = Settings.objects.create(content=content, type='title')
            newModel.save()
        
        return JsonResponse({
            'msg': 'success'
        })

def welcome_save(request):
    if request.method == 'POST':
        content = request.POST['content']

        if Settings.objects.filter(type='welcome').count() > 0:
            terms = Settings.objects.filter(type='welcome').get()
            terms.content = content
            terms.save()
        else:
            newModel = Settings.objects.create(content=content, type='welcome')
            newModel.save()
        
        return JsonResponse({
            'msg': 'success'
        })

def logoUpload(request):
    if request.method == "POST":
        form = SettingsForm(request.POST, request.FILES)
        if form.is_valid():
            if Settings_Image.objects.filter(type='logo').count() > 0:
                mm = Settings_Image.objects.filter(type='logo').get()
                mm.file.close()
                mm.file.delete()
                Settings_Image.objects.filter(type='logo').delete()
            
            file = request.FILES['file']
            form.save()
            img = Settings_Image.objects.last()
            img.name = file
            img.type = 'logo'
            img.save()
            return redirect('settings')
        else:
            return redirect('settings')

def loginUpload(request):
    if request.method == "POST":
        form = SettingsForm(request.POST, request.FILES)
        if form.is_valid():
            if Settings_Image.objects.filter(type='login').count() > 0:
                mm = Settings_Image.objects.filter(type='login').get()
                mm.file.close()
                mm.file.delete()
                Settings_Image.objects.filter(type='login').delete()

            file = request.FILES['file']
            form.save()
            img = Settings_Image.objects.last()
            img.name = file
            img.type = 'login'
            img.save()
            return redirect('settings')
        else:
            return redirect('settings')

def faviconUpload(request):
    if request.method == "POST":
        form = SettingsForm(request.POST, request.FILES)
        if form.is_valid():
            if Settings_Image.objects.filter(type='favicon').count() > 0:
                mm = Settings_Image.objects.filter(type='favicon').get()
                mm.file.close()
                mm.file.delete()
                Settings_Image.objects.filter(type='favicon').delete()

            file = request.FILES['file']
            form.save()
            img = Settings_Image.objects.last()
            img.name = file
            img.type = 'favicon'
            img.save()
            return redirect('settings')
        else:
            return redirect('settings')

@login_required
def users(request):
    user_list = User.objects.exclude(is_superuser=1)
    return render(request, 'users.html', {
        'user_list': user_list
    })
@login_required

def is_ajax(request):
    return request.META.get('HTTP_X_REQUESTED_WITH') == 'XMLHttpRequest'

@csrf_protect
def addUser(request):
    if is_ajax(request) and request.method == "POST":
        username = request.POST['username']
        pwd = request.POST['pwd']
        
        search_user = User.objects.filter(username=username).count()
        if search_user > 0 :
            return JsonResponse({
                'msg': 'user already exists'
            })

        newUser = User.objects.create(username=username, is_active=1)
        newUser.set_password(pwd)
        newUser.save()

        count = User.objects.all().count() - 1

        user_list = serializers.serialize('json', [newUser])
        return JsonResponse({
            'msg': 'success',
            'user_list': user_list,
            'count': count
        })

@csrf_protect
def operateUser(request):
    if is_ajax(request) and request.method == "POST":
        id = request.POST['id']
        type = request.POST['type']

        if type == 'delete':

            User.objects.filter(pk=id).delete()

            return JsonResponse({
                'msg': 'deleted'
            })
        
        if type == 'de-activate':
            user = User.objects.filter(pk=id).get()
            user.is_active = 0
            user.save()

            return JsonResponse({
                'msg': 'de-activated'
            })

        if type == 'activate':
            user = User.objects.filter(pk=id).get()
            user.is_active = 1
            user.save()

            return JsonResponse({
                'msg': 'activated'
            })

@csrf_protect
def editUser(request):
    if is_ajax(request) and request.method == "POST":
        id = request.POST['id']
        username = request.POST['username']
        pwd = request.POST['pwd']

        user = User.objects.filter(pk=id).get()
        user.username = username
        user.set_password(pwd)
        user.save()

        return JsonResponse({
                'msg': 'edited'
            })


@login_required
def master_sheet(request):
    sheets = MasterSheet.objects.all()
    # master_sheets = serializers.serialize('json', query_set)
    return render(request, 'input/master_sheet.html', {
        'sheets': sheets
    })
    

def uploadMasterSheet(request):
    if request.method == "POST":
        form = MasterSheetForm(request.POST, request.FILES)
        if form.is_valid():
            file = request.FILES['file']
            form.save()
            sheet = MasterSheet.objects.last()
            sheet.name = file
            sheet.save()
            # make other sheets inactive
            save_id = sheet.id
            other_sheets = MasterSheet.objects.exclude(id=save_id)
            for data in other_sheets:
                data.status = 0
                data.save() 

            return redirect('master_sheet')
        else:
            return redirect('master_sheet')

def operateMasterSheet(request):
    if is_ajax(request) and request.method == "POST":
        id = request.POST['id']
        type = request.POST['type']

        if type == 'delete':
            mm = MasterSheet.objects.filter(pk=id).get()
            mm.file.close()
            mm.file.delete()
            MasterSheet.objects.filter(pk=id).delete()

            return JsonResponse({
                'msg': 'deleted'
            })
        
        if type == 'de-activate':
            sheet = MasterSheet.objects.filter(pk=id).get()
            sheet.status = 0
            sheet.save()

            return JsonResponse({
                'msg': 'de-activated'
            })

        if type == 'activate':
            sheet = MasterSheet.objects.filter(pk=id).get()
            sheet.status = 1
            sheet.save()
            
            # make other sheets inactive
            save_id = sheet.id
            other_sheets = MasterSheet.objects.exclude(id=save_id)
            for data in other_sheets:
                data.status = 0
                data.save() 

            return JsonResponse({
                'msg': 'activated'
            })

def masterSheetDownload(request):
    if request.method == "POST":
        ids = request.POST['ids']
        ids = ids.split(',')

        if len(ids) > 0:

            files = MasterSheet.objects.filter(id__in=ids)

            paths = []
            names = []
            for file in files:
                prefix = os.getcwd() + '/media/master_sheets/'
                path = prefix + file.filename()
                paths.append(path)
                name = file.name
                names.append(name)
            
            print('path is ...', paths)

            if len(ids) == 1:
                file_path = paths[0]
                file_name = names[0]
                if os.path.exists(file_path):
                    with open(file_path, 'rb') as fh:
                        response = HttpResponse(fh.read(), content_type="application/vnd.ms-excel")
                        # response['Content-Disposition'] = 'inline; filename=' + os.path.basename(file_path)
                        response['Content-Disposition'] = 'inline; filename=' + file_name
                        return response
            else: 
                zip_filename = "master.zip"

                in_memory = BytesIO()
                zip = ZipFile(in_memory, "a")

                index = 0
                for path in paths:
                    fname = names[index]
                    zip.write(path, fname)
                    index += 1

                zip.close()
                
                resp = HttpResponse(content_type = "application/x-zip-compressed")
                resp['Content-Disposition'] = 'attachment; filename=%s' % zip_filename

                in_memory.seek(0)    
                resp.write(in_memory.read())

                return resp


@login_required
def read_sheet(request):
    sheets = ReadSheet.objects.all()
    return render(request, 'input/read_sheet.html', {
        'sheets': sheets
    })

def uploadReadSheet(request):
    if request.method == "POST":
        form = ReadSheetForm(request.POST, request.FILES)
        if form.is_valid():
            set_flag = 0
            if ReadSheet.objects.filter(status=1).count() > 0:
                origin_active_sheet = ReadSheet.objects.filter(status=1).get()
                set_flag = 1

            file = request.FILES['file']
            form.save()
            sheet = ReadSheet.objects.last()
            sheet.name = file
            sheet.save()

            # make other sheets inactive
            save_id = sheet.id
            other_sheets = ReadSheet.objects.exclude(id=save_id)
            for data in other_sheets:
                data.status = 0
                data.save() 

            # validate
            msg = ''

            read_path = os.getcwd() + '/media/read_sheets/' + ReadSheet.objects.filter(status=1).get().filename()
            if os.path.isfile(read_path):
                wb = openpyxl.load_workbook(read_path)
                ws = wb.active

                rows_cnt = ws.max_row
                cols_cnt = ws.max_column

                start_row = 0 #count real data starts
                flag = False
                for r in range(1, rows_cnt):
                    if flag == True:
                        break
                    for c in range(1, cols_cnt):
                        if ws.cell(row=r, column=c).value == "Q.No":
                            start_row = r + 1
                            flag = True
                            break
                
            for r in range(start_row, rows_cnt+1):
                if ws.cell(row=r, column=5).value == 'Q. picture single response' or ws.cell(row=r, column=5).value == 'Q. picture multiple response':
                    if ws.cell(row=r, column=3).value == None:
                        msg = '"Q. picture single response" or "Q. picture multiple response" is incorrect. please see if you have filled the "Question Media"field'  
                        # print(msg)

                if ws.cell(row=r, column=5).value == 'Q. media single response' or ws.cell(row=r, column=5).value == 'Q. media multiple response':
                    if ws.cell(row=r, column=3).value == None:
                        msg = '"Q. media single response" or "Q. media multiple response" is incorrect. please see if you have filled the "Question Media"field'  
                        # print(msg)

            if msg != '':
                mm = ReadSheet.objects.filter(pk=save_id).get()
                mm.file.close()
                mm.file.delete()
                ReadSheet.objects.filter(pk=save_id).delete()
                if set_flag == 1:
                    origin_active_sheet.status = 1
                    origin_active_sheet.save()
            
            sheets = ReadSheet.objects.all()
            return render(request, 'input/read_sheet.html', {
                'sheets': sheets,
                'msg': msg
            })           
            # return redirect('read_sheet')
        else:
            return redirect('read_sheet')

def operateReadSheet(request):
    if is_ajax(request) and request.method == "POST":
        id = request.POST['id']
        type = request.POST['type']

        if type == 'delete':
            mm = ReadSheet.objects.filter(pk=id).get()
            mm.file.close()
            mm.file.delete()
            ReadSheet.objects.filter(pk=id).delete()

            return JsonResponse({
                'msg': 'deleted'
            })
        
        if type == 'de-activate':
            sheet = ReadSheet.objects.filter(pk=id).get()
            sheet.status = 0
            sheet.save()

            return JsonResponse({
                'msg': 'de-activated'
            })

        if type == 'activate':
            sheet = ReadSheet.objects.filter(pk=id).get()
            sheet.status = 1
            sheet.save()
            
            # make other sheets inactive
            save_id = sheet.id
            other_sheets = ReadSheet.objects.exclude(id=save_id)
            for data in other_sheets:
                data.status = 0
                data.save() 

            return JsonResponse({
                'msg': 'activated'
            })

def readSheetDownload(request):
    if request.method == "POST":
        ids = request.POST['ids']
        ids = ids.split(',')

        print(ids)
        if len(ids) > 0:

            files = ReadSheet.objects.filter(id__in=ids)

            paths = []
            names = []
            for file in files:
                prefix = os.getcwd() + '/media/read_sheets/'
                path = prefix + file.filename()
                paths.append(path)
                name = file.name
                names.append(name)
            
            print('path is ...', paths)

            if len(ids) == 1:
                file_path = paths[0]
                file_name = names[0]
                if os.path.exists(file_path):
                    with open(file_path, 'rb') as fh:
                        response = HttpResponse(fh.read(), content_type="application/vnd.ms-excel")
                        # response['Content-Disposition'] = 'inline; filename=' + os.path.basename(file_path)
                        response['Content-Disposition'] = 'inline; filename=' + file_name
                        return response
            else:
                zip_filename = "read.zip"

                in_memory = BytesIO()
                zip = ZipFile(in_memory, "a")

                index = 0
                for path in paths:
                    # fname = os.path.split(path)[1]
                    fname = names[index]
                    zip.write(path, fname)
                    index += 1

                zip.close()
                
                resp = HttpResponse(content_type = "application/x-zip-compressed")
                resp['Content-Disposition'] = 'attachment; filename=%s' % zip_filename

                in_memory.seek(0)    
                resp.write(in_memory.read())

                return resp

@login_required
def interpretation_sheet(request):
    sheets = InterpretationSheet.objects.all()
    return render(request, 'input/interpretation_sheet.html', {
        'sheets': sheets
    })

def uploadInterpretationSheet(request):
    if request.method == "POST":
        form = InterpretationSheetForm(request.POST, request.FILES)
        if form.is_valid():
            file = request.FILES['file']
            form.save()
            sheet = InterpretationSheet.objects.last()
            sheet.name = file
            sheet.save()
            # make other sheets inactive
            save_id = sheet.id
            other_sheets = InterpretationSheet.objects.exclude(id=save_id)
            for data in other_sheets:
                data.status = 0
                data.save() 

            return redirect('interpretation_sheet')
        else:
            return redirect('interpretation_sheet')

def operateInterpretationSheet(request):
    if is_ajax(request) and request.method == "POST":
        id = request.POST['id']
        type = request.POST['type']

        if type == 'delete':
            mm = InterpretationSheet.objects.filter(pk=id).get()
            mm.file.close()
            mm.file.delete()
            InterpretationSheet.objects.filter(pk=id).delete()

            return JsonResponse({
                'msg': 'deleted'
            })
        
        if type == 'de-activate':
            sheet = InterpretationSheet.objects.filter(pk=id).get()
            sheet.status = 0
            sheet.save()

            return JsonResponse({
                'msg': 'de-activated'
            })

        if type == 'activate':
            sheet = InterpretationSheet.objects.filter(pk=id).get()
            sheet.status = 1
            sheet.save()
            
            # make other sheets inactive
            save_id = sheet.id
            other_sheets = InterpretationSheet.objects.exclude(id=save_id)
            for data in other_sheets:
                data.status = 0
                data.save() 

            return JsonResponse({
                'msg': 'activated'
            })

def interpretationSheetDownload(request):
    if request.method == "POST":
        ids = request.POST['ids']
        ids = ids.split(',')

        print(ids)
        if len(ids) > 0:

            files = InterpretationSheet.objects.filter(id__in=ids)

            paths = []
            names = []
            for file in files:
                prefix = os.getcwd() + '/media/interpretation_sheets/'
                path = prefix + file.filename()
                paths.append(path)
                name = file.name
                names.append(name)
            
            print('path is ...', paths)
            
            if len(ids) == 1:
                file_path = paths[0]
                file_name = names[0]
                if os.path.exists(file_path):
                    with open(file_path, 'rb') as fh:
                        response = HttpResponse(fh.read(), content_type="application/vnd.ms-excel")
                        # response['Content-Disposition'] = 'inline; filename=' + os.path.basename(file_path)
                        response['Content-Disposition'] = 'inline; filename=' + file_name
                        return response
            else:

                zip_filename = "interpretation.zip"

                in_memory = BytesIO()
                zip = ZipFile(in_memory, "a")

                index = 0
                for path in paths:
                    # fname = os.path.split(path)[1]
                    fname = names[index]
                    zip.write(path, fname)
                    index += 1

                zip.close()
                
                resp = HttpResponse(content_type = "application/x-zip-compressed")
                resp['Content-Disposition'] = 'attachment; filename=%s' % zip_filename

                in_memory.seek(0)    
                resp.write(in_memory.read())

                return resp

@login_required
def images(request):
    sheets = Images_Bot.objects.all()
    return render(request, 'attachments/images.html', {
        'sheets': sheets
    })

def uploadImages(request):
    if request.method == "POST":
        form = Images_BotForm(request.POST, request.FILES)
        if form.is_valid():
            file = request.FILES['file']
            form.save()
            sheet = Images_Bot.objects.last()
            sheet.name = file
            sheet.save()
            return redirect('images')
        else:
            return redirect('images')

def operateImages(request):
    if is_ajax(request) and request.method == "POST":
        id = request.POST['id']
        type = request.POST['type']

        if type == 'delete':
            mm = Images_Bot.objects.filter(pk=id).get()
            mm.file.close()
            mm.file.delete()
            Images_Bot.objects.filter(pk=id).delete()

            return JsonResponse({
                'msg': 'deleted'
            })
        
        if type == 'de-activate':
            sheet = Images_Bot.objects.filter(pk=id).get()
            sheet.status = 0
            sheet.save()

            return JsonResponse({
                'msg': 'de-activated'
            })

        if type == 'activate':
            sheet = Images_Bot.objects.filter(pk=id).get()
            sheet.status = 1
            sheet.save()

            return JsonResponse({
                'msg': 'activated'
            })

def imagesDownload(request):
    if request.method == "POST":
        ids = request.POST['ids']
        ids = ids.split(',')

        print(ids)
        if len(ids) > 0:

            files = Images_Bot.objects.filter(id__in=ids)

            paths = []
            names = []
            for file in files:
                prefix = os.getcwd() + '/media/attachments/images/'
                path = prefix + file.filename()
                paths.append(path)
                name = file.name
                names.append(name)
            
            print('path is ...', paths)

            if len(ids) == 1:
                file_path = paths[0]
                file_name = names[0]
                if os.path.exists(file_path):
                    with open(file_path, 'rb') as fh:
                        response = HttpResponse(fh.read(), content_type="application/vnd.ms-excel")
                        # response['Content-Disposition'] = 'inline; filename=' + os.path.basename(file_path)
                        response['Content-Disposition'] = 'inline; filename=' + file_name
                        return response
            else:
            
                zip_filename = "images.zip"

                in_memory = BytesIO()
                zip = ZipFile(in_memory, "a")

                index = 0
                for path in paths:
                    # fname = os.path.split(path)[1]
                    fname = names[index]
                    zip.write(path, fname)
                    index += 1

                zip.close()
                
                resp = HttpResponse(content_type = "application/x-zip-compressed")
                resp['Content-Disposition'] = 'attachment; filename=%s' % zip_filename

                in_memory.seek(0)    
                resp.write(in_memory.read())

                return resp

@login_required
def database(request):
    sheets = Database_Excel.objects.all()
    return render(request, 'attachments/database.html', {
        'sheets': sheets
    })

def uploadDatabase(request):
    if request.method == "POST":
        form = Database_ExcelForm(request.POST, request.FILES)
        if form.is_valid():
            file = request.FILES['file']
            form.save()
            sheet = Database_Excel.objects.last()
            sheet.name = file
            sheet.save()
            return redirect('database')
        else:
            return redirect('database')

def operateDatabase(request):
    if is_ajax(request) and request.method == "POST":
        id = request.POST['id']
        type = request.POST['type']

        if type == 'delete':
            mm = Database_Excel.objects.filter(pk=id).get()
            mm.file.close()
            mm.file.delete()
            Database_Excel.objects.filter(pk=id).delete()

            return JsonResponse({
                'msg': 'deleted'
            })
        
        if type == 'de-activate':
            sheet = Database_Excel.objects.filter(pk=id).get()
            sheet.status = 0
            sheet.save()

            return JsonResponse({
                'msg': 'de-activated'
            })

        if type == 'activate':
            sheet = Database_Excel.objects.filter(pk=id).get()
            sheet.status = 1
            sheet.save()

            return JsonResponse({
                'msg': 'activated'
            })

def databaseDownload(request):
    if request.method == "POST":
        ids = request.POST['ids']
        ids = ids.split(',')

        print(ids)
        if len(ids) > 0:

            files = Database_Excel.objects.filter(id__in=ids)

            paths = []
            names = []
            for file in files:
                prefix = os.getcwd() + '/media/attachments/database/'
                path = prefix + file.filename()
                paths.append(path)
                name = file.name
                names.append(name)
            
            print('path is ...', paths)

            if len(ids) == 1:
                file_path = paths[0]
                file_name = names[0]
                if os.path.exists(file_path):
                    with open(file_path, 'rb') as fh:
                        response = HttpResponse(fh.read(), content_type="application/vnd.ms-excel")
                        # response['Content-Disposition'] = 'inline; filename=' + os.path.basename(file_path)
                        response['Content-Disposition'] = 'inline; filename=' + file_name
                        return response
            else:
            
                zip_filename = "database.zip"

                in_memory = BytesIO()
                zip = ZipFile(in_memory, "a")

                index = 0
                for path in paths:
                    # fname = os.path.split(path)[1]
                    fname = names[index]
                    zip.write(path, fname)
                    index += 1

                zip.close()
                
                resp = HttpResponse(content_type = "application/x-zip-compressed")
                resp['Content-Disposition'] = 'attachment; filename=%s' % zip_filename

                in_memory.seek(0)    
                resp.write(in_memory.read())

                return resp

@login_required
def write_sheet(request):
    read_path = os.path.dirname(os.path.dirname(os.path.abspath(__file__))) + "/media/write_sheets/write.xlsx"

    if os.path.exists(read_path):
        wb = openpyxl.load_workbook(read_path)
        ws = wb.active

        rows_cnt = ws.max_row
        cols_cnt = ws.max_column
        
        total_result = []
        for r in range(1, rows_cnt + 1):
            record = []
            for i in range(1, cols_cnt + 1):
                record.append(ws.cell(row=r, column=i).value)
            total_result.append(record)

        thead = total_result[0]
        total_result.pop(0)

        return render(request, 'output/write_sheet.html', {
            'thead': thead,
            'dataset': total_result,
            'btn_display': True
        })
    else:
        return render(request, 'output/write_sheet.html', {
            'btn_display': False
        })

def exportWriteSheet(request):
    if request.method == 'POST':
        BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))

        file_path = BASE_DIR + '/media/write_sheets/write.xlsx'
        if os.path.exists(file_path):
            with open(file_path, 'rb') as fh:
                response = HttpResponse(fh.read(), content_type="application/vnd.ms-excel")
                response['Content-Disposition'] = 'inline; filename=' + os.path.basename(file_path)
                return response
        else:
            return render(request, 'output/write_sheet.html')

def deleteWriteSheet(request):
    path = os.getcwd() + '/media/write_sheets/write.xlsx'
    if os.path.isfile(path):
        os.remove(path)
        return JsonResponse({
            'msg': 'success'
        })

@login_required
def aa_output_sheet(request):
    read_path =  os.path.dirname(os.path.dirname(os.path.abspath(__file__))) + "/media/aa_outputs/aaOutputSheet.xlsx"
    
    if os.path.exists(read_path):
        wb = openpyxl.load_workbook(read_path)
        ws = wb.active

        rows_cnt = ws.max_row
        cols_cnt = ws.max_column

        total_result = []
        for r in range(1, rows_cnt + 1):
            record = []
            for i in range(1, cols_cnt + 1):
                record.append(ws.cell(row=r, column=i).value)
            total_result.append(record)

        thead = total_result[0]
        total_result.pop(0)
        
        return render(request, 'output/aa_output.html', {
            'thead': thead,
            'dataset': total_result,
            'btn_display': True
        })
    else:
        return render(request, 'output/aa_output.html', {
            'btn_display': False
        })
        
def exportAAOutputSheet(request):
    if request.method == 'POST':
        BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))

        file_path = BASE_DIR + '/media/aa_outputs/aaOutputSheet.xlsx'
        
        if os.path.exists(file_path):
            with open(file_path, 'rb') as fh:
                response = HttpResponse(fh.read(), content_type="application/vnd.ms-excel")
                response['Content-Disposition'] = 'inline; filename=' + os.path.basename(file_path)
                return response
        else:
            return render(request, 'output/aa_output.html')

def deleteAAOutputSheet(request):
    path = os.getcwd() + '/media/aa_outputs/aaOutputSheet.xlsx'
    if os.path.isfile(path):
        os.remove(path)
        return JsonResponse({
            'msg': 'success'
        })

@login_required
def videos(request):
    return render(request, 'attachments/videos.html')


@login_required
def chatbot(request):
    msg = ''
    if InterpretationSheet.objects.count() == 0:
        msg = 'Interpretation sheet does not exist!'
    else:
        if InterpretationSheet.objects.filter(status=1).count() == 0:
            msg = 'Interpretation sheet is inactive!'

    if MasterSheet.objects.count() == 0:
        msg = 'Master sheet does not exist!'
    else:
        if MasterSheet.objects.filter(status=1).count() == 0: 
            msg = 'Master sheet is inactive!'
    
    if ReadSheet.objects.count() == 0:
        msg = 'Read sheet does not exist!'
    else:
        if ReadSheet.objects.filter(status=1).count() == 0:
            msg = 'Read sheet is inactive!'

    if msg == '':
        read_path = os.getcwd() + '/media/read_sheets/' + ReadSheet.objects.filter(status=1).get().filename()
        
        if os.path.isfile(read_path):
            wb = openpyxl.load_workbook(read_path)
            ws = wb.active

            rows_cnt = ws.max_row
            cols_cnt = ws.max_column

            start_row = 0 #count real data starts
            flag = False
            for r in range(1, rows_cnt):
                if flag == True:
                    break
                for c in range(1, cols_cnt):
                    if ws.cell(row=r, column=c).value == "Q.No":
                        start_row = r + 1
                        flag = True
                        break

            for r in range(start_row, rows_cnt+1):
                if (ws.cell(row=r, column=5).value) == 'Database response':
                    db = ws.cell(row=r, column=7).value

                    if Database_Excel.objects.filter(name=db).filter(status=1).count() == 0:
                        msg = 'Database ' + db + ' does not exist or inactive. please check this problem!'

                if (ws.cell(row=r, column=5).value) == 'Q. picture single response' or (ws.cell(row=r, column=5).value) == 'Q. picture multiple response':
                    img = ws.cell(row=r, column=3).value

                    if Images_Bot.objects.filter(name=img).filter(status=1).count() == 0:
                        msg = 'Image ' + img + ' does not exist or inactive. please check this problem!'
                
                if (ws.cell(row=r, column=5).value) == 'Single picture response' or (ws.cell(row=r, column=5).value) == 'Multiple picture response':
                    img_list = []
                    count = 0
                    for k in range(r+1, rows_cnt+1):
                        count = count + 1
                        if ws.cell(row=k, column=5).value != None:
                            break
                    for t in range(0, count):
                        img_list.append(ws.cell(row=r+t, column=7).value.split('>')[0])
                        
                    if len(img_list):
                        for img in img_list:
                            if Images_Bot.objects.filter(name=img).filter(status=1).count() == 0:
                                msg = 'Image ' + img + ' does not exist or inactive. please check this problem!'

    return render(request, 'chatbot.html', {
        'msg': msg
    })

def chatbot_start(request):
    if ReadSheet.objects.count() > 0:
        if ReadSheet.objects.filter(status=1).count() > 0:
            read_path = os.getcwd() + "/media/read_sheets/" + ReadSheet.objects.filter(status=1).get().filename()
    # print('read path is ....................', read_path)

    global dataset
    dataset = getData(read_path)
    # print('data test------------', dataset[6])
    start_data = dataset[0]
    return JsonResponse({
        'start_data': start_data 
    })

def changePassword(request):
    return render(request, "change_password.html")

def checkOldPassword(request):
    if request.method == 'POST':
        oldPass = request.POST['oldPassword']
        dbPass = User.objects.filter(is_superuser=1).get().password
        
        check = check_password(oldPass, dbPass)
       
        if check == True:
            return JsonResponse({
                'msg': 'success'
            }) 
        else:
            return JsonResponse({
                'msg': 'fail'
            })

def amendPass(request):
    if request.method == 'POST':
        new_pass = request.POST['new_pass']

        superadmin = User.objects.filter(is_superuser=1).get()
        superadmin.set_password(new_pass)
        superadmin.save()
        return JsonResponse({
            'msg': 'success'
        })



def getNextQuery(request):
    if request.method == 'POST':
        index = int(request.POST['index'])

        data = dataset[index - 1]

        return JsonResponse({
            'dataset': data
        })

def chatbot_getDatabase(request):
    if request.method == 'POST':
        filename = request.POST['filename']

        data = getDatabase(filename)
        return JsonResponse({
            'dataset': data
        })

def writeOutput(request):
    if request.method == "POST":
        write_dataset = json.loads(request.POST['write_dataset'])

        username = None
        # if request.user.is_authenticated():
        username = request.user.username
        if ReadSheet.objects.count() > 0:
            if ReadSheet.objects.filter(status=1).count() > 0:
                read_path = os.getcwd() + "/media/read_sheets/" + ReadSheet.objects.filter(status=1).get().filename()

        write_path = os.getcwd() + "/media/write_sheets/write.xlsx"
        writeExcel(write_dataset, read_path, write_path, username)

        if MasterSheet.objects.count() > 0:
            if MasterSheet.objects.filter(status=1).count() > 0:
                master_path = os.getcwd() + "/media/master_sheets/" + MasterSheet.objects.filter(status=1).get().filename()

        data_science(master_path, write_path)

        return JsonResponse({
            'msg': 'file processing is succeed!'
        })

def get_Feedback(request):
    if request.method == 'POST':
        mode = request.POST['mode']
        
        if InterpretationSheet.objects.count() > 0:
            if InterpretationSheet.objects.filter(status=1).count() > 0:
                interpretation_path = os.getcwd() + "/media/interpretation_sheets/" + InterpretationSheet.objects.filter(status=1).get().filename()

        feedbacks = getFeedback(mode, interpretation_path)

        return JsonResponse({
            'feedbacks': feedbacks
        })


def test_sheets(request):
    sheets = TestSheet.objects.all()
    return render(request, 'test_sheets.html', {
        'sheets': sheets
    })

def upload_test_sheets(request):
    if request.method == "POST":
        form = TestSheetForm(request.POST, request.FILES)
        if form.is_valid():
            file = request.FILES['file']
            form.save()
            sheet = TestSheet.objects.last()
            sheet.name = file
            sheet.save()
            return redirect('test_sheets')
        else:
            return redirect('test_sheets')

def processTestSheet(request):
    if is_ajax(request) and request.method == "POST":
        if request.method == 'POST':
            id = request.POST['id']
            file = TestSheet.objects.filter(pk=id).get()
            file_path = file.file
            aa_name = file.name

            if MasterSheet.objects.count() > 0:
                if MasterSheet.objects.filter(status=1).count() > 0:
                    master_path = os.getcwd() + "/media/master_sheets/" + MasterSheet.objects.filter(status=1).get().filename()

            write_path =  os.getcwd() + '/media/' + str(file_path)
            test_science(master_path, write_path, aa_name)

            return JsonResponse({
                'msg': 'success',
                'id': id
            })

            

def downloadAATestSheet(request):
    if request.method == 'POST':
        id = request.POST['id']
        file_name = TestSheet.objects.filter(pk=id).get().name
        file_path = "media/aa_tests/AAoutput_" + file_name
        print(file_path)
        if os.path.exists(file_path):
            with open(file_path, 'rb') as fh:
                response = HttpResponse(fh.read(), content_type="application/vnd.ms-excel")
                response['Content-Disposition'] = 'inline; filename=' + os.path.basename(file_path)
                return response
        else:
            return redirect('test_sheets')

def downloadTestSheet(request):
    if request.method == 'POST':
        BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))

        id = request.POST['id']
        file = TestSheet.objects.filter(pk=id).get()
        name = file.name
        path = str(file.file)
        file_path = BASE_DIR + '/media/' + path
        
        if os.path.exists(file_path):
            with open(file_path, 'rb') as fh:
                response = HttpResponse(fh.read(), content_type="application/vnd.ms-excel")
                response['Content-Disposition'] = 'inline; filename=' + os.path.basename(name)
                return response
        else:
            return redirect('test_sheets')

def operateTestSheet(request):
    if is_ajax(request) and request.method == "POST":
        id = request.POST['id']
        type = request.POST['type']

        if type == 'delete':
            mm = TestSheet.objects.filter(pk=id).get()
            mm.file.close()
            mm.file.delete()
            TestSheet.objects.filter(pk=id).delete()

            return JsonResponse({
                'msg': 'deleted'
            })

# @login_required
# def terms_conditions(request):
#     return render(request, 'terms_and_conditions.html')

# @login_required
# def cookie_policy(request):
#     return render(request, 'cookie_policy.html')


# def checkAdmin(request):
#     if request.user.is_superuser == 1:
#         return True
#     else:
#         return render(request, '404.html')